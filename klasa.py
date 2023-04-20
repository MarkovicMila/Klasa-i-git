import openpyxl as openpyxl
import psycopg2 as psycopg2

class Ljudi:
    def __init__(self):
        self.upit=""
        self.sql_result=None

    def kreiraj_upit(self,upit):
        self.upit=upit
    
    def get_sql(self):
        try:
            con=psycopg2.connect(
                database='ljudi',
                user='postgres',
                port='5432',
                host='localhost',
                password='itoip'
            )
            cursor=con.cursor()
            cursor.execute(self.upit)
            self.sql_result=cursor.fetchall()

        except(Exception,psycopg2.Error) as e:
            print('Error: ',e)

        finally:
            con.close()
            cursor.close()
    
    def export_excel(self,naziv):
        if self.sql_result!=None:
            wb=openpyxl.Workbook()
            ws=wb.active
            ws.title=naziv
            ws['A1'].value='JMBG'
            ws['B1'].value='Ime'
            ws['C1'].value='Prezime'
            ws['D1'].value='Godine'
            ws['E1'].value='Pol'
            for i in range(2,len(self.sql_result)+2):
                ws.cell(row=i,column=1).value=self.sql_result[i-2][0]
                ws.cell(row=i,column=2).value=self.sql_result[i-2][1]
                ws.cell(row=i,column=3).value=self.sql_result[i-2][2]
                ws.cell(row=i,column=4).value=self.sql_result[i-2][3]
                ws.cell(row=i,column=5).value=self.sql_result[i-2][4]
            
            wb.save(filename='{}.xlsx'.format(naziv))
            wb.close()
            return 'Excel file kreiran'
    def dodaj_coveka(self,jmbg,ime,prezime,godine,pol):
        try:
            con=psycopg2.connect(
                database='ljudi',
                user='postgres',
                host='localhost',
                port='5432',
                password='itoip'
            )
            cursor=con.cursor()
            com='''INSERT INTO LJUDI VALUES ('{}','{}','{}',{},'{}');'''.format(jmbg,ime,prezime,int(godine),pol)
            cursor.execute(com)
            con.commit()
        except(Exception,psycopg2.Error) as e:
            print('Error: ',e)
        finally:
            con.close()
            cursor.close()
L=Ljudi()